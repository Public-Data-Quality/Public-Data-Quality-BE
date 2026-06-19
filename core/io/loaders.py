from __future__ import annotations

import csv
from collections.abc import Iterator
from functools import lru_cache
from pathlib import Path

from ..config.constants import (
    UPLOAD_DATASET_ID_PREFIX,
    UPLOAD_DATASET_TYPE,
    UPLOAD_PROVIDER_CODE,
    UPLOAD_PROVIDER_NAME,
    UPLOAD_SERVICE_TYPE,
    UPLOAD_UPDATE_CYCLE,
)
from ..schema.models import DatasetMeta, StandardTerm


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_headers(values: list[object]) -> list[str]:
    return [_stringify_cell(value) for value in values]


def _iter_csv_rows(path: Path) -> Iterator[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            yield {str(key): (value or "") for key, value in row.items() if key is not None}


def _iter_xlsx_rows(path: Path) -> Iterator[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLSX 업로드를 처리하려면 openpyxl이 필요합니다.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    headers = _clean_headers(list(header_row or []))
    if not any(headers):
        workbook.close()
        raise ValueError("Uploaded dataset has no header row.")

    try:
        for row in row_iter:
            values = list(row or [])
            yield {
                header: _stringify_cell(values[index]) if index < len(values) else ""
                for index, header in enumerate(headers)
                if header
            }
    finally:
        workbook.close()


def _iter_xls_rows(path: Path) -> Iterator[dict[str, str]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLS 업로드를 처리하려면 xlrd가 필요합니다.") from exc

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("Uploaded dataset has no header row.")

    headers = _clean_headers(sheet.row_values(0))
    if not any(headers):
        raise ValueError("Uploaded dataset has no header row.")

    for row_index in range(1, sheet.nrows):
        values = sheet.row_values(row_index)
        yield {
            header: _stringify_cell(values[index]) if index < len(values) else ""
            for index, header in enumerate(headers)
            if header
        }


def iter_uploaded_rows(file_path: str | Path) -> Iterator[dict[str, str]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _iter_csv_rows(path)
        return
    if suffix == ".xlsx":
        yield from _iter_xlsx_rows(path)
        return
    if suffix == ".xls":
        yield from _iter_xls_rows(path)
        return
    raise ValueError(f"Unsupported file type: {suffix or '<none>'}. Supported: .csv, .xlsx, .xls")


def load_uploaded_headers(file_path: str | Path) -> list[str]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            return [header.strip() for header in next(reader, []) if str(header).strip()]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ValueError("XLSX 업로드를 처리하려면 openpyxl이 필요합니다.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(values_only=True), None)
        workbook.close()
        return [header for header in _clean_headers(list(header_row or [])) if header]
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover
            raise ValueError("XLS 업로드를 처리하려면 xlrd가 필요합니다.") from exc
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            return []
        return [header for header in _clean_headers(sheet.row_values(0)) if header]
    raise ValueError(f"Unsupported file type: {suffix or '<none>'}. Supported: .csv, .xlsx, .xls")


def _split_csv_list(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def _file_cache_key(file_path: str | Path) -> tuple[str, int]:
    path = Path(file_path).resolve()
    return str(path), path.stat().st_mtime_ns


def load_standard_terms(csv_path: str | Path) -> tuple[dict[str, StandardTerm], dict[str, str]]:
    terms, synonyms = _load_standard_terms_cached(*_file_cache_key(csv_path))
    return dict(terms), dict(synonyms)


@lru_cache(maxsize=4)
def _load_standard_terms_cached(path_key: str, _mtime_ns: int) -> tuple[dict[str, StandardTerm], dict[str, str]]:
    path = Path(path_key)
    terms: dict[str, StandardTerm] = {}
    synonyms: dict[str, str] = {}

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            name = (row.get("공통표준용어명") or "").strip()
            if not name:
                continue
            synonym_list = _split_csv_list(row.get("용어 이음동의어 목록", ""))
            term = StandardTerm(
                name=name,
                description=(row.get("공통표준용어설명") or "").strip(),
                abbreviation=(row.get("공통표준용어영문약어명") or "").strip(),
                domain_name=(row.get("공통표준도메인명") or "").strip(),
                allowed_values=(row.get("허용값") or "").strip(),
                storage_format=(row.get("저장 형식") or "").strip(),
                expression_format=(row.get("표현 형식") or "").strip(),
                code_name=(row.get("행정표준코드명") or "").strip(),
                owner_org=(row.get("소관기관명") or "").strip(),
                synonyms=synonym_list,
            )
            terms[name] = term
            synonyms[name] = name
            for synonym in synonym_list:
                synonyms[synonym] = name

    return terms, synonyms


def load_dataset_meta(
    csv_path: str | Path,
    dataset_id: str | None = None,
    dataset_name: str | None = None,
) -> DatasetMeta:
    path = Path(csv_path)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            row_id = (row.get("목록키") or "").strip()
            row_name = (row.get("목록명") or "").strip()
            if dataset_id and row_id != dataset_id:
                continue
            if dataset_name and row_name != dataset_name:
                continue
            total_rows = (row.get("전체행") or "").strip()
            return DatasetMeta(
                dataset_id=row_id,
                dataset_name=row_name,
                keywords=_split_csv_list(row.get("키워드", "")),
                provider_name=(row.get("제공기관명") or "").strip(),
                provider_code=(row.get("제공기관코드") or "").strip(),
                dataset_type=(row.get("목록유형") or "").strip(),
                service_type=(row.get("서비스 유형") or "").strip(),
                data_format=(row.get("데이터포맷") or "").strip(),
                request_fields=_split_csv_list(row.get("요청변수", "")),
                response_fields=_split_csv_list(row.get("출력결과", "")),
                update_cycle=(row.get("주기") or "").strip(),
                total_rows=int(total_rows) if total_rows.isdigit() else None,
            )
    target = dataset_id or dataset_name or "<unknown>"
    raise ValueError(f"Dataset not found: {target}")


def load_uploaded_dataset_meta(file_path: str | Path, dataset_name: str | None = None) -> DatasetMeta:
    path = Path(file_path)
    suffix = path.suffix.lstrip(".").lower() or "csv"
    header = load_uploaded_headers(path)
    if not header:
        raise ValueError("Uploaded dataset has no header row.")

    name = dataset_name or path.stem
    return DatasetMeta(
        dataset_id=f"{UPLOAD_DATASET_ID_PREFIX}{path.stem}",
        dataset_name=name,
        keywords=[],
        provider_name=UPLOAD_PROVIDER_NAME,
        provider_code=UPLOAD_PROVIDER_CODE,
        dataset_type=UPLOAD_DATASET_TYPE,
        service_type=UPLOAD_SERVICE_TYPE,
        data_format=suffix,
        request_fields=[],
        response_fields=[column.strip() for column in header if column.strip()],
        update_cycle=UPLOAD_UPDATE_CYCLE,
        total_rows=None,
    )


def build_example_index(meta_csv_path: str | Path, limit: int = 5000) -> dict[str, list[str]]:
    examples = _build_example_index_cached(*_file_cache_key(meta_csv_path), limit)
    return {key: list(values) for key, values in examples.items()}


@lru_cache(maxsize=4)
def _build_example_index_cached(path_key: str, _mtime_ns: int, limit: int = 5000) -> dict[str, list[str]]:
    path = Path(path_key)
    examples: dict[str, list[str]] = {}
    seen = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            dataset_name = (row.get("목록명") or "").strip()
            for field in _split_csv_list(row.get("출력결과", "")):
                bucket = examples.setdefault(field, [])
                if len(bucket) < 5:
                    bucket.append(dataset_name)
            seen += 1
            if seen >= limit:
                break
    return examples
